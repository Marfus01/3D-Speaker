#!/usr/bin/env python3
"""Collect the four EXP-02 active-speaker result groups from saved artifacts.

The script deliberately reuses the existing Stage-4 timestamps and face
embeddings.  It never reruns TalkNet, face tracking, FQA, Stage 5, or the HMM.
For each saved active visual-speaker face (AVSF), it first verifies the stored
bbox against the saved embedding.  Only mismatches fall back to MTCNN on that
single saved frame, followed by CurricularFace embedding matching.

Typical server commands (run from ``speaker-diarization``)::

    python local/analyze_exp02_active_speaker.py --task summarize
    python local/analyze_exp02_active_speaker.py --task recover-bbox --device cuda
    python local/analyze_exp02_active_speaker.py --task plot-bbox
    python local/analyze_exp02_active_speaker.py --task candidates

Omitting ``--tv-name`` processes both configured TV datasets.  All outputs are
written under ``runs/<tv_name>/exp_video/active_speaker_results`` by default.
"""

from __future__ import annotations

import argparse
import bisect
import csv
import hashlib
import json
import math
import pickle
import re
import shutil
import subprocess
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, time as datetime_time
from pathlib import Path
from typing import Any, Iterable, Sequence

import numpy as np
import pandas as pd


LOCAL_DIR = Path(__file__).resolve().parent
RECIPE_ROOT = LOCAL_DIR.parent
REPO_ROOT = RECIPE_ROOT.parents[2]
if str(RECIPE_ROOT) not in sys.path:
    sys.path.insert(0, str(RECIPE_ROOT))
if str(LOCAL_DIR) not in sys.path:
    sys.path.insert(0, str(LOCAL_DIR))

from speakerlab.utils.config import yaml_config_loader  # noqa: E402
from utils_metric import (  # noqa: E402
    class_matching,
    main_character_list_BB,
    main_character_list_IL,
)


DATASETS = {
    "the big bang theory": {
        "short_name": "BigBang",
        "annotation_name": "text_annotated_BB.xlsx",
        "main_characters": main_character_list_BB,
    },
    "I love my family": {
        "short_name": "LoveFamily",
        "annotation_name": "text_annotated_IL.xlsx",
        "main_characters": main_character_list_IL,
    },
}

RESULT_COLUMNS = [
    "dataset",
    "episode",
    "record_index",
    "time",
    "frame_idx",
    "segment_id",
    "stored_x1",
    "stored_y1",
    "stored_x2",
    "stored_y2",
    "x1",
    "y1",
    "x2",
    "y2",
    "recovery_method",
    "recovery_status",
    "n_detected_faces",
    "top_cosine",
    "second_cosine",
    "cosine_margin",
]

MANUAL_REVIEW_COLUMNS = [
    "final_label",
    "review_status",
    "selected_for_paper",
    "scenario_tags",
    "review_note",
]

FINAL_LABELS = ["TP", "FP", "FN", "TN", "Ambiguous", "Exclude"]

REFERENCE_IDENTITY_PERCENT = {
    "the big bang theory": 93.81,
    "I love my family": 94.80,
}


@dataclass(frozen=True)
class DatasetPaths:
    tv_name: str
    short_name: str
    repo_root: Path
    data_root: Path
    recipe_root: Path
    exp_dir: Path
    output_dir: Path
    embs_dir: Path
    result_dir: Path
    subseg_json: Path
    annotation: Path
    video_list: Path
    conf_file: Path
    onnx_dir: Path

    @property
    def data_dir(self) -> Path:
        return self.output_dir / "data"

    @property
    def plots_dir(self) -> Path:
        return self.output_dir / "plots"

    @property
    def media_dir(self) -> Path:
        return self.output_dir / "media"

    @property
    def review_dir(self) -> Path:
        return self.output_dir / "review"

    @property
    def recommendations_dir(self) -> Path:
        return self.output_dir / "recommendations"


class SegmentIndex:
    """Resolve an episode-local timestamp to a subtitle segment ID."""

    def __init__(self, subsegments: dict[str, dict[str, Any]]) -> None:
        grouped: dict[str, list[tuple[float, float, str]]] = defaultdict(list)
        for segment_id, info in subsegments.items():
            episode = segment_id.rsplit("-", 1)[0]
            grouped[episode].append(
                (float(info["start"]), float(info["stop"]), segment_id)
            )
        self._intervals: dict[str, list[tuple[float, float, str]]] = {}
        self._starts: dict[str, list[float]] = {}
        for episode, intervals in grouped.items():
            intervals.sort(key=lambda item: (item[0], item[1], item[2]))
            self._intervals[episode] = intervals
            self._starts[episode] = [item[0] for item in intervals]

    def find(self, episode: str, timestamp: float) -> str:
        intervals = self._intervals.get(episode, [])
        starts = self._starts.get(episode, [])
        if not intervals:
            return ""
        index = bisect.bisect_right(starts, timestamp)
        matches = [
            segment_id
            for start, stop, segment_id in intervals[max(0, index - 4) : index + 1]
            if start <= timestamp < stop
        ]
        return matches[0] if len(matches) == 1 else ""


class VideoFrameReader:
    """Keep one OpenCV capture open while seeking episode frames."""

    def __init__(self, video_path: Path) -> None:
        import cv2

        self._cv2 = cv2
        self.video_path = video_path
        self.capture = cv2.VideoCapture(str(video_path))
        if not self.capture.isOpened():
            raise RuntimeError(f"Cannot open video: {video_path}")
        self.fps = float(self.capture.get(cv2.CAP_PROP_FPS))
        if not math.isclose(self.fps, 25.0, rel_tol=0.0, abs_tol=1e-3):
            raise ValueError(f"Expected a 25-fps Stage-4 video, got {self.fps}: {video_path}")

    def read(self, frame_idx: int) -> np.ndarray:
        self.capture.set(self._cv2.CAP_PROP_POS_FRAMES, int(frame_idx))
        ok, frame = self.capture.read()
        if not ok:
            raise RuntimeError(f"Cannot read frame {frame_idx} from {self.video_path}")
        return frame

    def close(self) -> None:
        self.capture.release()


class BboxRecoverer:
    """Recover the bbox whose crop produced a saved AVSF embedding."""

    def __init__(
        self,
        conf_file: Path,
        onnx_dir: Path,
        device: str,
        device_id: int,
        stored_bbox_min_cosine: float,
        candidate_min_cosine: float,
        candidate_min_margin: float,
    ) -> None:
        import cv2
        import torch
        from facenet_pytorch import MTCNN
        from vision_tools.face_recognition import FaceRecIR101

        conf = yaml_config_loader(str(conf_file))
        if device == "cuda" and not torch.cuda.is_available():
            print("[WARNING] CUDA is unavailable; bbox recovery falls back to CPU.")
            device = "cpu"
        self.cv2 = cv2
        self.torch_device = torch.device(f"cuda:{device_id}" if device == "cuda" else "cpu")
        self.face_detector = MTCNN(
            image_size=160,
            margin=0,
            min_face_size=int(conf["min_face_size"]),
            thresholds=[0.6, 0.7, 0.7],
            factor=0.709,
            post_process=True,
            device=self.torch_device,
            keep_all=True,
        )
        self.face_recognizer = FaceRecIR101(
            str(onnx_dir), device=device, device_id=device_id
        )
        self.min_box_size = float(conf["min_box_size"])
        self.min_box_prob = float(conf["min_box_prob"])
        self.stored_bbox_min_cosine = stored_bbox_min_cosine
        self.candidate_min_cosine = candidate_min_cosine
        self.candidate_min_margin = candidate_min_margin

    @staticmethod
    def _clip_bbox(bbox: Sequence[float], frame: np.ndarray) -> np.ndarray:
        height, width = frame.shape[:2]
        return np.asarray(
            [
                max(int(bbox[0]), 0),
                max(int(bbox[1]), 0),
                min(int(bbox[2]), width),
                min(int(bbox[3]), height),
            ],
            dtype=np.int32,
        )

    def _embedding(self, frame: np.ndarray, bbox: Sequence[float]) -> np.ndarray | None:
        clipped = self._clip_bbox(bbox, frame)
        x1, y1, x2, y2 = clipped.tolist()
        if x2 <= x1 or y2 <= y1:
            return None
        crop = frame[y1:y2, x1:x2]
        if crop.size == 0:
            return None
        return np.asarray(self.face_recognizer(crop), dtype=np.float32).reshape(-1)

    @staticmethod
    def _cosine(left: np.ndarray | None, right: np.ndarray) -> float:
        if left is None:
            return float("nan")
        left = np.asarray(left, dtype=np.float32).reshape(-1)
        right = np.asarray(right, dtype=np.float32).reshape(-1)
        denom = float(np.linalg.norm(left) * np.linalg.norm(right))
        return float(np.dot(left, right) / denom) if denom > 0 else float("nan")

    def _detect(self, frame: np.ndarray) -> list[np.ndarray]:
        rgb = self.cv2.cvtColor(frame, self.cv2.COLOR_BGR2RGB)
        boxes, probs = self.face_detector.detect(rgb)
        if boxes is None or probs is None:
            return []
        valid: list[tuple[float, np.ndarray]] = []
        for box, prob in zip(boxes, probs):
            clipped = self._clip_bbox(np.maximum(box, 0), frame)
            width = int(clipped[2] - clipped[0])
            height = int(clipped[3] - clipped[1])
            if min(width, height) >= self.min_box_size and float(prob) >= self.min_box_prob:
                valid.append((float(prob), clipped))
        valid.sort(key=lambda item: item[0], reverse=True)
        return [bbox for _, bbox in valid]

    def recover(
        self,
        frame: np.ndarray,
        stored_bbox: Sequence[float],
        saved_embedding: np.ndarray,
    ) -> dict[str, Any]:
        stored_bbox = self._clip_bbox(stored_bbox, frame)
        stored_embedding = self._embedding(frame, stored_bbox)
        stored_cosine = self._cosine(stored_embedding, saved_embedding)
        if np.isfinite(stored_cosine) and stored_cosine >= self.stored_bbox_min_cosine:
            return {
                "bbox": stored_bbox,
                "method": "stored_bbox_embedding_match",
                "status": "resolved",
                "n_detected_faces": 0,
                "top_cosine": stored_cosine,
                "second_cosine": float("nan"),
                "cosine_margin": float("inf"),
            }

        boxes = self._detect(frame)
        scored: list[tuple[float, np.ndarray]] = []
        for bbox in boxes:
            candidate_embedding = self._embedding(frame, bbox)
            cosine = self._cosine(candidate_embedding, saved_embedding)
            if np.isfinite(cosine):
                scored.append((cosine, bbox))
        scored.sort(key=lambda item: item[0], reverse=True)
        if not scored:
            return {
                "bbox": np.full(4, np.nan),
                "method": "mtcnn_embedding_match",
                "status": "unresolved_no_face",
                "n_detected_faces": len(boxes),
                "top_cosine": float("nan"),
                "second_cosine": float("nan"),
                "cosine_margin": float("nan"),
            }
        top_cosine, top_bbox = scored[0]
        second_cosine = scored[1][0] if len(scored) > 1 else float("nan")
        margin = top_cosine - second_cosine if np.isfinite(second_cosine) else float("inf")
        resolved = (
            top_cosine >= self.candidate_min_cosine
            and margin >= self.candidate_min_margin
        )
        return {
            "bbox": top_bbox if resolved else np.full(4, np.nan),
            "method": "mtcnn_embedding_match",
            "status": "resolved" if resolved else "unresolved_low_confidence",
            "n_detected_faces": len(boxes),
            "top_cosine": top_cosine,
            "second_cosine": second_cosine,
            "cosine_margin": margin,
        }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--task",
        choices=["all", "summarize", "recover-bbox", "plot-bbox", "candidates"],
        default="all",
    )
    parser.add_argument(
        "--tv-name",
        action="append",
        choices=list(DATASETS),
        help="Repeat to process multiple datasets; default: both datasets.",
    )
    parser.add_argument("--repo-root", type=Path, default=REPO_ROOT)
    parser.add_argument(
        "--data-root",
        type=Path,
        default=Path("/data/home/scv7387/run/tv_series_plus/dataset"),
    )
    parser.add_argument("--exp-dir", type=Path, help="Allowed only with one --tv-name.")
    parser.add_argument("--output-dir", type=Path, help="Allowed only with one --tv-name.")
    parser.add_argument("--annotation", type=Path, help="Allowed only with one --tv-name.")
    parser.add_argument("--video-list", type=Path, help="Allowed only with one --tv-name.")
    parser.add_argument("--conf", type=Path, help="Allowed only with one --tv-name.")
    parser.add_argument("--onnx-dir", type=Path, help="Allowed only with one --tv-name.")
    parser.add_argument("--device", choices=["cpu", "cuda"], default="cuda")
    parser.add_argument("--device-id", type=int, default=0)
    parser.add_argument("--stored-bbox-min-cosine", type=float, default=0.999)
    parser.add_argument("--candidate-min-cosine", type=float, default=0.95)
    parser.add_argument("--candidate-min-margin", type=float, default=0.05)
    parser.add_argument("--checkpoint-every", type=int, default=500)
    parser.add_argument("--max-records", type=int, help="Smoke-test limit per dataset.")
    parser.add_argument("--candidates-per-class", type=int, default=20)
    parser.add_argument("--min-recovery-rate", type=float, default=0.99)
    parser.add_argument(
        "--require-reference-identity",
        action="store_true",
        help="Stop unless identity agreement rounds to the registered paper value.",
    )
    clip_group = parser.add_mutually_exclusive_group()
    clip_group.add_argument(
        "--write-clips",
        dest="write_clips",
        action="store_true",
        help="Write deduplicated full-frame clips with audio (default).",
    )
    clip_group.add_argument(
        "--skip-clips",
        dest="write_clips",
        action="store_false",
        help="Build images/workbook without encoding clips.",
    )
    parser.set_defaults(write_clips=True)
    parser.add_argument(
        "--clip-context",
        type=float,
        default=1.0,
        help="Seconds added before and after each subtitle/event clip.",
    )
    parser.add_argument(
        "--event-max-gap",
        type=float,
        default=1.0,
        help="Maximum gap for grouping AVSF records without a subtitle segment.",
    )
    parser.add_argument("--ffmpeg", default="ffmpeg")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()
    selected = args.tv_name or list(DATASETS)
    if len(selected) != 1 and any(
        value is not None
        for value in [args.exp_dir, args.output_dir, args.annotation, args.video_list, args.conf, args.onnx_dir]
    ):
        parser.error("Dataset-specific path overrides require exactly one --tv-name.")
    if args.checkpoint_every <= 0 or args.candidates_per_class <= 0:
        parser.error("--checkpoint-every and --candidates-per-class must be positive.")
    if not 0 < args.min_recovery_rate <= 1:
        parser.error("--min-recovery-rate must be in (0, 1].")
    if args.clip_context < 0 or args.event_max_gap <= 0:
        parser.error("--clip-context must be non-negative and --event-max-gap positive.")
    args.tv_name = selected
    return args


def _first_existing(candidates: Iterable[Path]) -> Path:
    candidates = list(candidates)
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    return candidates[0].resolve()


def resolve_paths(args: argparse.Namespace, tv_name: str) -> DatasetPaths:
    repo_root = args.repo_root.resolve()
    recipe_root = repo_root / "egs/3dspeaker/speaker-diarization"
    exp_dir = (args.exp_dir or recipe_root / "runs" / tv_name / "exp_video").resolve()
    data_tv_dir = args.data_root.resolve() / tv_name
    annotation = args.annotation or _first_existing(
        [
            data_tv_dir / "annotation/text_annotated.xlsx",
            repo_root.parent / "annotations" / DATASETS[tv_name]["annotation_name"],
        ]
    )
    return DatasetPaths(
        tv_name=tv_name,
        short_name=DATASETS[tv_name]["short_name"],
        repo_root=repo_root,
        data_root=args.data_root.resolve(),
        recipe_root=recipe_root,
        exp_dir=exp_dir,
        output_dir=(args.output_dir or exp_dir / "active_speaker_results").resolve(),
        embs_dir=exp_dir / "embs_video",
        result_dir=exp_dir / "result",
        subseg_json=exp_dir / "json/subseg.json",
        annotation=Path(annotation).resolve(),
        video_list=(args.video_list or data_tv_dir / "raw/video.list").resolve(),
        conf_file=(args.conf or recipe_root / "conf" / tv_name / "diar_video.yaml").resolve(),
        onnx_dir=(args.onnx_dir or recipe_root / "pretrained_models").resolve(),
    )


def require_files(paths: Iterable[Path], purpose: str) -> None:
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise FileNotFoundError(f"Missing {purpose} file(s):\n  " + "\n  ".join(missing))


def data_artifact(paths: DatasetPaths, filename: str) -> Path:
    """Prefer the structured data directory, with legacy-root read support."""
    current = paths.data_dir / filename
    legacy = paths.output_dir / filename
    return current if current.is_file() or not legacy.is_file() else legacy


def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as stream:
        return json.load(stream)


def atomic_json_dump(value: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as stream:
        json.dump(value, stream, ensure_ascii=False, indent=2, allow_nan=False)
    temporary.replace(path)


def atomic_pickle_dump(value: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("wb") as stream:
        pickle.dump(value, stream)
    temporary.replace(path)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_manifest(paths: DatasetPaths, task: str, inputs: Iterable[Path]) -> None:
    entries = []
    for path in sorted(set(Path(item).resolve() for item in inputs)):
        entries.append(
            {
                "path": str(path),
                "exists": path.is_file(),
                "sha256": sha256(path) if path.is_file() else None,
            }
        )
    manifest = {
        "created_at": datetime.now().astimezone().isoformat(),
        "task": task,
        "tv_name": paths.tv_name,
        "repo_root": str(paths.repo_root),
        "data_root": str(paths.data_root),
        "exp_dir": str(paths.exp_dir),
        "output_dir": str(paths.output_dir),
        "inputs": entries,
    }
    atomic_json_dump(manifest, paths.output_dir / "run_manifest.json")


def _time_to_seconds(value: Any) -> float:
    if isinstance(value, datetime_time):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1e6
    if isinstance(value, datetime):
        return _time_to_seconds(value.time())
    if isinstance(value, (int, float, np.number)) and not pd.isna(value):
        return float(value) * 86400.0 if 0 <= float(value) < 1 else float(value)
    hours, minutes, seconds = map(float, str(value).strip().split(":"))
    return hours * 3600 + minutes * 60 + seconds


def _visible_faces(row: pd.Series) -> list[str]:
    faces: list[str] = []
    for column in ["frontal face", "Side face"]:
        value = row.get(column)
        if pd.isna(value):
            continue
        faces.extend(item.strip() for item in str(value).split(",") if item.strip())
    return sorted(set(faces))


def load_annotations(paths: DatasetPaths) -> pd.DataFrame:
    require_files([paths.annotation], "annotation")
    frame = pd.read_excel(paths.annotation)
    required = {
        "Episode",
        "Text Index",
        "Start Time",
        "End Time",
        "whether annotate speaker",
        "speaker",
        "whether annotate face",
        "frontal face",
        "Side face",
    }
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"Annotation is missing columns {missing}: {paths.annotation}")
    frame = frame[frame["whether annotate speaker"].astype(str).str.strip() == "Yes"].copy()
    frame["segment_id"] = frame.apply(
        lambda row: f"E{int(row['Episode']):02}-{int(row['Text Index'])}", axis=1
    )
    frame["start_seconds"] = frame["Start Time"].map(_time_to_seconds)
    frame["stop_seconds"] = frame["End Time"].map(_time_to_seconds)
    frame["mid_seconds"] = (frame["start_seconds"] + frame["stop_seconds"]) / 2.0
    frame["visible_faces"] = frame.apply(_visible_faces, axis=1)
    frame["speaker_face_visible"] = frame.apply(
        lambda row: str(row["speaker"]).strip() in row["visible_faces"], axis=1
    )
    return frame


def _load_cluster_files(paths: DatasetPaths) -> tuple[dict[str, int], dict[str, int]]:
    unique_path = paths.result_dir / "cluster_results_vision_vad_uniq.json"
    majority_path = paths.result_dir / "cluster_results_vision_vad_major.json"
    require_files([unique_path, majority_path], "visual cluster result")
    unique = {key: int(value) for key, value in read_json(unique_path).items()}
    majority = {key: int(value) for key, value in read_json(majority_path).items()}
    return unique, majority


def _normalize_speaker(speaker: Any, main_characters: Sequence[str]) -> str:
    speaker = str(speaker).strip()
    return speaker if speaker in main_characters else "Others"


def build_identity_records(
    paths: DatasetPaths, annotations: pd.DataFrame, unique_labels: dict[str, int]
) -> tuple[pd.DataFrame, dict[int, str]]:
    frame = annotations[annotations["segment_id"].isin(unique_labels)].copy()
    if frame.empty:
        raise ValueError(f"No annotated segments overlap unique visual labels for {paths.tv_name}.")
    frame["cluster_label"] = frame["segment_id"].map(unique_labels).astype(int)
    main_characters = DATASETS[paths.tv_name]["main_characters"]
    frame["annotated_speaker"] = frame["speaker"].map(
        lambda value: _normalize_speaker(value, main_characters)
    )

    nonnegative = frame[frame["cluster_label"] >= 0]
    original_clusters = sorted(nonnegative["cluster_label"].unique().tolist())
    cluster_to_index = {label: index for index, label in enumerate(original_clusters)}
    speakers = sorted(set(frame["annotated_speaker"].tolist() + ["Others"]))
    speaker_to_index = {speaker: index for index, speaker in enumerate(speakers)}
    mapping: dict[int, str] = {-1: "Others"}
    if not nonnegative.empty:
        reference = np.eye(len(speakers), dtype=int)[
            nonnegative["annotated_speaker"].map(speaker_to_index).to_numpy()
        ]
        prediction = np.eye(len(original_clusters), dtype=int)[
            nonnegative["cluster_label"].map(cluster_to_index).to_numpy()
        ]
        mapped_indices = class_matching(
            reference, prediction, others_chara_id=speaker_to_index["Others"]
        )
        mapping.update(
            {
                original_clusters[cluster_index]: speakers[speaker_index]
                for cluster_index, speaker_index in mapped_indices.items()
            }
        )
    frame["mapped_speaker"] = frame["cluster_label"].map(mapping).fillna("Others")
    frame["is_correct"] = frame["mapped_speaker"] == frame["annotated_speaker"]

    cluster_counts = frame.groupby(["cluster_label", "annotated_speaker"]).size()
    totals = frame.groupby("cluster_label").size()
    purities = {
        int(cluster): float(cluster_counts.loc[cluster].max() / totals.loc[cluster])
        for cluster in totals.index
    }
    frame["cluster_purity"] = frame["cluster_label"].map(purities)
    return frame, mapping


def summarize_dataset(paths: DatasetPaths) -> None:
    unique_labels, majority_labels = _load_cluster_files(paths)
    require_files([paths.subseg_json], "subsegment")
    subsegments = read_json(paths.subseg_json)
    annotations = load_annotations(paths)
    identity, mapping = build_identity_records(paths, annotations, unique_labels)
    paths.data_dir.mkdir(parents=True, exist_ok=True)

    coverage = pd.DataFrame(
        [
            {
                "dataset": paths.short_name,
                "N_seg": len(subsegments),
                "N_any": len(majority_labels),
                "C_any": len(majority_labels) / len(subsegments),
                "N_uniq": len(unique_labels),
                "C_uniq": len(unique_labels) / len(subsegments),
            }
        ]
    )
    coverage.to_csv(paths.data_dir / "coverage_summary.csv", index=False)

    identity_columns = [
        "segment_id",
        "Episode",
        "Text Index",
        "start_seconds",
        "stop_seconds",
        "mid_seconds",
        "speaker",
        "annotated_speaker",
        "cluster_label",
        "mapped_speaker",
        "is_correct",
        "cluster_purity",
        "whether annotate face",
        "visible_faces",
        "speaker_face_visible",
    ]
    identity[identity_columns].to_csv(
        paths.data_dir / "identity_agreement_records.csv", index=False
    )
    correct = int(identity["is_correct"].sum())
    summary = pd.DataFrame(
        [
            {
                "dataset": paths.short_name,
                "annotated_unique_segments": len(identity),
                "correct": correct,
                "incorrect": len(identity) - correct,
                "oracle_hungarian_accuracy": correct / len(identity),
            }
        ]
    )
    summary.to_csv(paths.data_dir / "identity_agreement_summary.csv", index=False)

    matrix = pd.crosstab(identity["cluster_label"], identity["annotated_speaker"])
    matrix.to_csv(paths.data_dir / "cluster_speaker_matrix.csv")
    atomic_json_dump(
        {str(cluster): speaker for cluster, speaker in sorted(mapping.items())},
        paths.data_dir / "cluster_to_speaker_mapping.json",
    )
    write_manifest(
        paths,
        "summarize",
        [
            paths.subseg_json,
            paths.annotation,
            paths.result_dir / "cluster_results_vision_vad_uniq.json",
            paths.result_dir / "cluster_results_vision_vad_major.json",
        ],
    )
    print(
        f"[INFO] {paths.short_name}: coverage={len(unique_labels)}/{len(subsegments)}, "
        f"identity agreement={correct}/{len(identity)}={correct / len(identity):.4f}"
    )


def require_reference_identity(paths: DatasetPaths) -> None:
    summary_path = data_artifact(paths, "identity_agreement_summary.csv")
    require_files([summary_path], "identity agreement summary")
    summary = pd.read_csv(summary_path)
    if len(summary) != 1 or "oracle_hungarian_accuracy" not in summary:
        raise ValueError(f"Unexpected identity summary format: {summary_path}")
    actual_percent = round(float(summary.iloc[0]["oracle_hungarian_accuracy"]) * 100, 2)
    expected_percent = REFERENCE_IDENTITY_PERCENT[paths.tv_name]
    if actual_percent != expected_percent:
        raise RuntimeError(
            f"{paths.short_name} identity agreement is {actual_percent:.2f}%, "
            f"expected {expected_percent:.2f}%; stopping before bbox recovery."
        )
    print(
        f"[GATE PASS] {paths.short_name}: identity agreement "
        f"{actual_percent:.2f}% matches the registered result."
    )


def read_video_list(path: Path) -> dict[str, Path]:
    require_files([path], "video list")
    mapping: dict[str, Path] = {}
    with path.open("r", encoding="utf-8") as stream:
        for line in stream:
            text = line.strip()
            if not text:
                continue
            video = Path(text)
            mapping[video.stem] = video
    if not mapping:
        raise ValueError(f"No videos found in {path}")
    return mapping


def _bbox_rows_to_pickle(rows: Sequence[dict[str, Any]]) -> dict[str, np.ndarray]:
    def numbers(names: Sequence[str], dtype: Any) -> np.ndarray:
        return np.asarray([[row[name] for name in names] for row in rows], dtype=dtype)

    return {
        "dataset": np.asarray([row["dataset"] for row in rows]),
        "episode": np.asarray([row["episode"] for row in rows]),
        "record_index": np.asarray([row["record_index"] for row in rows], dtype=np.int64),
        "time": np.asarray([row["time"] for row in rows], dtype=np.float64),
        "frame_idx": np.asarray([row["frame_idx"] for row in rows], dtype=np.int64),
        "segment_id": np.asarray([row["segment_id"] for row in rows]),
        "stored_bbox": numbers(["stored_x1", "stored_y1", "stored_x2", "stored_y2"], np.float64),
        "bbox": numbers(["x1", "y1", "x2", "y2"], np.float64),
        "recovery_method": np.asarray([row["recovery_method"] for row in rows]),
        "recovery_status": np.asarray([row["recovery_status"] for row in rows]),
        "n_detected_faces": np.asarray([row["n_detected_faces"] for row in rows], dtype=np.int64),
        "top_cosine": np.asarray([row["top_cosine"] for row in rows], dtype=np.float64),
        "second_cosine": np.asarray([row["second_cosine"] for row in rows], dtype=np.float64),
        "cosine_margin": np.asarray([row["cosine_margin"] for row in rows], dtype=np.float64),
    }


def save_bbox_rows(rows: Sequence[dict[str, Any]], output_dir: Path) -> None:
    rows = sorted(rows, key=lambda row: (row["episode"], row["record_index"]))
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "active_face_bbox_records.csv"
    temporary = csv_path.with_suffix(".csv.tmp")
    with temporary.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=RESULT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(csv_path)
    # This pickle is the canonical per-AVSF bbox artifact used by plot-bbox.
    atomic_pickle_dump(
        _bbox_rows_to_pickle(rows), output_dir / "active_face_bboxes.pkl"
    )


def load_bbox_rows(output_dir: Path) -> list[dict[str, Any]]:
    csv_path = output_dir / "active_face_bbox_records.csv"
    if not csv_path.is_file():
        return []
    frame = pd.read_csv(csv_path)
    return frame.to_dict("records")


def recover_bboxes(paths: DatasetPaths, args: argparse.Namespace) -> None:
    require_files([paths.subseg_json, paths.conf_file], "bbox recovery")
    if not (paths.onnx_dir / "face_recog_ir101.onnx").is_file():
        raise FileNotFoundError(
            f"Missing CurricularFace model: {paths.onnx_dir / 'face_recog_ir101.onnx'}"
        )
    vad_files = sorted(paths.embs_dir.glob("*_vad.pkl"))
    if not vad_files:
        raise FileNotFoundError(f"No *_vad.pkl files in {paths.embs_dir}")
    video_paths = read_video_list(paths.video_list)
    segment_index = SegmentIndex(read_json(paths.subseg_json))
    bbox_output_dir = paths.data_dir
    existing_rows = [] if args.overwrite else load_bbox_rows(
        data_artifact(paths, "active_face_bbox_records.csv").parent
    )
    completed = {
        (str(row["episode"]), int(row["record_index"])) for row in existing_rows
    }
    rows = list(existing_rows)
    recoverer = BboxRecoverer(
        paths.conf_file,
        paths.onnx_dir,
        args.device,
        args.device_id,
        args.stored_bbox_min_cosine,
        args.candidate_min_cosine,
        args.candidate_min_margin,
    )
    processed_this_run = 0
    stop = False
    for vad_file in vad_files:
        episode = vad_file.name[: -len("_vad.pkl")]
        if episode not in video_paths:
            raise KeyError(f"{episode} is missing from {paths.video_list}")
        with vad_file.open("rb") as stream:
            saved = pickle.load(stream)
        required_keys = {"embeddings", "times", "bboxes"}
        if set(saved) != required_keys:
            raise ValueError(f"Unexpected schema in {vad_file}: {sorted(saved)}")
        embeddings = np.asarray(saved["embeddings"])
        times = np.asarray(saved["times"], dtype=float)
        stored_bboxes = np.asarray(saved["bboxes"])
        if not (len(embeddings) == len(times) == len(stored_bboxes)):
            raise ValueError(f"Length mismatch in {vad_file}")
        reader = VideoFrameReader(video_paths[episode])
        try:
            for record_index, (timestamp, stored_bbox, embedding) in enumerate(
                zip(times, stored_bboxes, embeddings)
            ):
                key = (episode, record_index)
                if key in completed:
                    continue
                if args.max_records is not None and processed_this_run >= args.max_records:
                    stop = True
                    break
                frame_idx = int(round(float(timestamp) * reader.fps))
                try:
                    frame = reader.read(frame_idx)
                    recovered = recoverer.recover(frame, stored_bbox, embedding)
                except Exception as error:  # Preserve the AVSF row and make failure auditable.
                    recovered = {
                        "bbox": np.full(4, np.nan),
                        "method": "error",
                        "status": f"unresolved_error:{type(error).__name__}",
                        "n_detected_faces": 0,
                        "top_cosine": float("nan"),
                        "second_cosine": float("nan"),
                        "cosine_margin": float("nan"),
                    }
                bbox = np.asarray(recovered["bbox"], dtype=float)
                stored_bbox = np.asarray(stored_bbox, dtype=float)
                row = {
                    "dataset": paths.short_name,
                    "episode": episode,
                    "record_index": record_index,
                    "time": float(timestamp),
                    "frame_idx": frame_idx,
                    "segment_id": segment_index.find(episode, float(timestamp)),
                    "stored_x1": stored_bbox[0],
                    "stored_y1": stored_bbox[1],
                    "stored_x2": stored_bbox[2],
                    "stored_y2": stored_bbox[3],
                    "x1": bbox[0],
                    "y1": bbox[1],
                    "x2": bbox[2],
                    "y2": bbox[3],
                    "recovery_method": recovered["method"],
                    "recovery_status": recovered["status"],
                    "n_detected_faces": recovered["n_detected_faces"],
                    "top_cosine": recovered["top_cosine"],
                    "second_cosine": recovered["second_cosine"],
                    "cosine_margin": recovered["cosine_margin"],
                }
                rows.append(row)
                completed.add(key)
                processed_this_run += 1
                if processed_this_run % args.checkpoint_every == 0:
                    save_bbox_rows(rows, bbox_output_dir)
                    print(
                        f"[INFO] {paths.short_name}: recovered/checkpointed "
                        f"{processed_this_run} new AVSF records",
                        flush=True,
                    )
        finally:
            reader.close()
        if stop:
            break
    save_bbox_rows(rows, bbox_output_dir)
    resolved = sum(row["recovery_status"] == "resolved" for row in rows)
    print(
        f"[INFO] {paths.short_name}: saved {len(rows)} AVSF bbox rows; "
        f"resolved={resolved}, unresolved={len(rows) - resolved}"
    )
    write_manifest(
        paths,
        "recover-bbox",
        [paths.subseg_json, paths.conf_file, paths.video_list, *vad_files],
    )


def plot_bbox_histogram(paths: DatasetPaths, min_recovery_rate: float) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    bbox_path = data_artifact(paths, "active_face_bboxes.pkl")
    require_files([bbox_path], "recovered bbox")
    with bbox_path.open("rb") as stream:
        records = pickle.load(stream)
    bboxes = np.asarray(records["bbox"], dtype=float)
    status = np.asarray(records["recovery_status"]).astype(str)
    valid = (status == "resolved") & np.isfinite(bboxes).all(axis=1)
    vad_files = sorted(paths.embs_dir.glob("*_vad.pkl"))
    if not vad_files:
        raise FileNotFoundError(f"No *_vad.pkl files in {paths.embs_dir}")
    expected_records = 0
    for vad_file in vad_files:
        with vad_file.open("rb") as stream:
            expected_records += len(pickle.load(stream)["times"])
    recovery_rate = float(valid.sum() / expected_records) if expected_records else 0.0
    if len(status) != expected_records:
        print(
            f"[WARNING] Recovered bbox artifact contains {len(status)} rows, but "
            f"the saved Stage-4 files contain {expected_records} AVSF records."
        )
    if recovery_rate < min_recovery_rate:
        raise RuntimeError(
            f"Resolved AVSF bbox rate {recovery_rate:.2%} is below the required "
            f"{min_recovery_rate:.2%}; refusing to generate a formal histogram."
        )
    bboxes = bboxes[valid]
    if len(bboxes) == 0:
        raise ValueError(f"No resolved AVSF bboxes in {bbox_path}")
    areas = np.maximum(0, bboxes[:, 2] - bboxes[:, 0]) * np.maximum(
        0, bboxes[:, 3] - bboxes[:, 1]
    )
    displayed_areas = np.minimum(areas, 100000)

    fig, axis = plt.subplots(figsize=(6.4, 4.4))
    axis.hist(displayed_areas, bins=np.linspace(0, 100000, 51), color="#1f77b4")
    for boundary in [10000, 20000, 30000, 40000]:
        axis.axvline(boundary, color="red", linestyle="--", linewidth=1.2)
    axis.set_title(f"Number of Pixels of Active Speaker Faces in {paths.tv_name} Test Set")
    axis.set_xlabel("Number of Pixels")
    axis.set_ylabel("Count")
    axis.set_xlim(0, 120000)
    ticks = [0, 20000, 40000, 60000, 80000, 100000]
    axis.set_xticks(ticks, ["0", "20000", "40000", "60000", "80000", ">=100000"])
    fig.tight_layout()
    paths.plots_dir.mkdir(parents=True, exist_ok=True)
    for suffix in ["pdf", "png"]:
        fig.savefig(
            paths.plots_dir / f"active_face_bbox_histogram.{suffix}",
            dpi=300,
            bbox_inches="tight",
        )
    plt.close(fig)
    print(
        f"[INFO] {paths.short_name}: plotted {len(areas)} resolved AVSF bboxes "
        f"from {bbox_path}"
    )


def load_identity_outputs(paths: DatasetPaths) -> pd.DataFrame:
    path = data_artifact(paths, "identity_agreement_records.csv")
    require_files([path], "identity agreement")
    frame = pd.read_csv(path)
    for column in ["is_correct", "speaker_face_visible"]:
        if column in frame:
            frame[column] = frame[column].astype(str).str.lower().map({"true": True, "false": False})
    return frame


def _clean_text(value: Any) -> str:
    return "" if pd.isna(value) else str(value).strip()


def _safe_token(value: Any) -> str:
    token = re.sub(r"[^A-Za-z0-9_.-]+", "_", _clean_text(value))
    return token.strip("_.") or "unknown"


def _face_names(value: Any) -> list[str]:
    if pd.isna(value):
        return []
    return [item.strip() for item in str(value).split(",") if item.strip()]


def _face_orientation(annotation: pd.Series | None, speaker: str) -> str:
    if annotation is None or not speaker:
        return ""
    raw_speaker = _clean_text(annotation.get("speaker"))
    aliases = {speaker, raw_speaker}
    frontal = set(_face_names(annotation.get("frontal face")))
    side = set(_face_names(annotation.get("Side face")))
    in_frontal = bool(aliases & frontal)
    in_side = bool(aliases & side)
    if in_frontal and in_side:
        return "frontal+side"
    if in_frontal:
        return "frontal"
    if in_side:
        return "side"
    return ""


def _bbox_area_bin(area: float) -> str:
    if not np.isfinite(area):
        return ""
    for upper in [10000, 20000, 30000, 40000]:
        if area <= upper:
            return f"<={upper}"
    return ">40000"


def _candidate_pool(
    paths: DatasetPaths,
    annotations: pd.DataFrame,
    identity: pd.DataFrame,
    bbox_frame: pd.DataFrame,
    subsegments: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    annotations_by_segment = {
        str(row["segment_id"]): row for _, row in annotations.iterrows()
    }
    identity_by_segment = {
        str(row["segment_id"]): row for _, row in identity.iterrows()
    }
    active_segments = {
        _clean_text(value) for value in bbox_frame["segment_id"] if _clean_text(value)
    }
    rows: list[dict[str, Any]] = []

    for _, bbox in bbox_frame.iterrows():
        episode = _clean_text(bbox["episode"])
        record_index = int(bbox["record_index"])
        frame_idx = int(bbox["frame_idx"])
        segment_id = _clean_text(bbox.get("segment_id"))
        annotation = annotations_by_segment.get(segment_id)
        identity_row = identity_by_segment.get(segment_id)
        segment = subsegments.get(segment_id, {})
        annotated_speaker = (
            _clean_text(identity_row.get("annotated_speaker"))
            if identity_row is not None
            else _normalize_speaker(
                annotation.get("speaker"), DATASETS[paths.tv_name]["main_characters"]
            )
            if annotation is not None
            else ""
        )
        mapped_speaker = (
            _clean_text(identity_row.get("mapped_speaker"))
            if identity_row is not None
            else ""
        )
        if identity_row is not None:
            auto_label = "TP" if bool(identity_row["is_correct"]) else "FP"
            auto_reason = (
                "mapped speaker matches annotation"
                if auto_label == "TP"
                else "mapped speaker differs from annotation"
            )
        else:
            auto_label = "UNSCORED"
            auto_reason = "no annotated unique-label identity result for this AVSF"
        coordinates = [float(bbox[column]) for column in ["x1", "y1", "x2", "y2"]]
        area = (
            max(0.0, coordinates[2] - coordinates[0])
            * max(0.0, coordinates[3] - coordinates[1])
            if np.isfinite(coordinates).all()
            else float("nan")
        )
        visible_faces = annotation["visible_faces"] if annotation is not None else []
        start = (
            float(annotation["start_seconds"])
            if annotation is not None
            else float(segment.get("start", bbox["time"]))
        )
        stop = (
            float(annotation["stop_seconds"])
            if annotation is not None
            else float(segment.get("stop", bbox["time"]))
        )
        rows.append(
            {
                "candidate_id": (
                    f"{_safe_token(episode)}_AVSF_{record_index:06d}_F{frame_idx:07d}"
                ),
                "record_type": "AVSF",
                "dataset": paths.short_name,
                "episode": episode,
                "segment_id": segment_id,
                "record_index": record_index,
                "time": float(bbox["time"]),
                "frame_idx": frame_idx,
                "start_seconds": start,
                "stop_seconds": stop,
                "annotated_speaker": annotated_speaker,
                "mapped_speaker": mapped_speaker,
                "cluster_label": (
                    int(identity_row["cluster_label"])
                    if identity_row is not None
                    else ""
                ),
                "cluster_purity": (
                    float(identity_row["cluster_purity"])
                    if identity_row is not None
                    else float("nan")
                ),
                "speaker_face_visible": (
                    bool(annotation["speaker_face_visible"])
                    if annotation is not None
                    else ""
                ),
                "visible_faces": ", ".join(visible_faces),
                "n_visible_faces": len(visible_faces),
                "face_orientation": _face_orientation(annotation, annotated_speaker),
                "x1": coordinates[0],
                "y1": coordinates[1],
                "x2": coordinates[2],
                "y2": coordinates[3],
                "bbox_area": area,
                "bbox_area_bin": _bbox_area_bin(area),
                "frame_width": "",
                "frame_height": "",
                "face_position": "",
                "recovery_method": _clean_text(bbox.get("recovery_method")),
                "recovery_status": _clean_text(bbox.get("recovery_status")),
                "top_cosine": bbox.get("top_cosine", float("nan")),
                "cosine_margin": bbox.get("cosine_margin", float("nan")),
                "auto_label": auto_label,
                "auto_reason": auto_reason,
            }
        )

    for _, annotation in annotations.iterrows():
        segment_id = str(annotation["segment_id"])
        if segment_id in active_segments:
            continue
        visible_faces = annotation["visible_faces"]
        if bool(annotation["speaker_face_visible"]):
            auto_label = "FN"
            auto_reason = "annotated speaker face visible but no saved AVSF"
        elif visible_faces:
            auto_label = "TN"
            auto_reason = "only non-speaker faces visible and no saved AVSF"
        else:
            auto_label = "UNSCORED"
            auto_reason = "no saved AVSF and no annotated visible face"
        episode = f"E{int(annotation['Episode']):02}"
        frame_time = float(annotation["mid_seconds"])
        annotated_speaker = _normalize_speaker(
            annotation["speaker"], DATASETS[paths.tv_name]["main_characters"]
        )
        rows.append(
            {
                "candidate_id": f"{episode}_SEG_{_safe_token(segment_id)}_MID",
                "record_type": "NO_ACTIVE_SEGMENT",
                "dataset": paths.short_name,
                "episode": episode,
                "segment_id": segment_id,
                "record_index": "",
                "time": frame_time,
                "frame_idx": int(round(frame_time * 25)),
                "start_seconds": float(annotation["start_seconds"]),
                "stop_seconds": float(annotation["stop_seconds"]),
                "annotated_speaker": annotated_speaker,
                "mapped_speaker": "",
                "cluster_label": "",
                "cluster_purity": float("nan"),
                "speaker_face_visible": bool(annotation["speaker_face_visible"]),
                "visible_faces": ", ".join(visible_faces),
                "n_visible_faces": len(visible_faces),
                "face_orientation": _face_orientation(annotation, annotated_speaker),
                "x1": float("nan"),
                "y1": float("nan"),
                "x2": float("nan"),
                "y2": float("nan"),
                "bbox_area": float("nan"),
                "bbox_area_bin": "",
                "frame_width": "",
                "frame_height": "",
                "face_position": "",
                "recovery_method": "",
                "recovery_status": "not_applicable",
                "top_cosine": float("nan"),
                "cosine_margin": float("nan"),
                "auto_label": auto_label,
                "auto_reason": auto_reason,
            }
        )

    pool = pd.DataFrame(rows)
    if pool.empty:
        raise ValueError(f"No qualitative candidates could be generated for {paths.tv_name}")
    if pool["candidate_id"].duplicated().any():
        duplicated = pool.loc[pool["candidate_id"].duplicated(), "candidate_id"].tolist()
        raise ValueError(f"Duplicate candidate IDs: {duplicated[:5]}")
    return pool.sort_values(["episode", "time", "record_type", "candidate_id"]).reset_index(
        drop=True
    )


def _face_position(row: pd.Series) -> str:
    if not all(np.isfinite(float(row[column])) for column in ["x1", "y1", "x2", "y2"]):
        return ""
    width = float(row["frame_width"])
    height = float(row["frame_height"])
    if width <= 0 or height <= 0:
        return ""
    center_x = (float(row["x1"]) + float(row["x2"])) / 2.0 / width
    center_y = (float(row["y1"]) + float(row["y2"])) / 2.0 / height
    horizontal = "left" if center_x < 1 / 3 else "right" if center_x > 2 / 3 else "center"
    vertical = "upper" if center_y < 1 / 3 else "lower" if center_y > 2 / 3 else "middle"
    return f"{horizontal}-{vertical}"


def _draw_full_frame(frame: np.ndarray, row: pd.Series, output_path: Path) -> None:
    import cv2

    canvas = frame.copy()
    colors = {
        "TP": (40, 170, 40),
        "FP": (40, 40, 220),
        "FN": (0, 140, 255),
        "TN": (220, 120, 30),
        "UNSCORED": (180, 180, 180),
    }
    label = str(row["auto_label"])
    color = colors.get(label, colors["UNSCORED"])
    if row["record_type"] == "AVSF" and all(
        np.isfinite(float(row[column])) for column in ["x1", "y1", "x2", "y2"]
    ):
        cv2.rectangle(
            canvas,
            (int(row["x1"]), int(row["y1"])),
            (int(row["x2"]), int(row["y2"])),
            color,
            4,
        )
    text = f"{label} suggestion | {row['candidate_id']}"
    cv2.rectangle(canvas, (0, 0), (min(canvas.shape[1], 1100), 42), (0, 0, 0), -1)
    cv2.putText(canvas, text, (12, 29), cv2.FONT_HERSHEY_SIMPLEX, 0.68, color, 2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not cv2.imwrite(str(output_path), canvas):
        raise RuntimeError(f"Failed to save full frame: {output_path}")


def _write_face_crop(frame: np.ndarray, row: pd.Series, output_path: Path) -> None:
    import cv2

    height, width = frame.shape[:2]
    x1 = max(0, min(width, int(row["x1"])))
    y1 = max(0, min(height, int(row["y1"])))
    x2 = max(0, min(width, int(row["x2"])))
    y2 = max(0, min(height, int(row["y2"])))
    if x2 <= x1 or y2 <= y1:
        raise ValueError(f"Invalid recovered bbox for {row['candidate_id']}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not cv2.imwrite(str(output_path), frame[y1:y2, x1:x2]):
        raise RuntimeError(f"Failed to save face crop: {output_path}")


def _write_clip(
    video_path: Path,
    output_path: Path,
    ffmpeg: str,
    start: float,
    stop: float,
) -> None:
    start = max(0.0, float(start))
    duration = max(0.04, float(stop) - start)
    wav_path = video_path.with_suffix(".wav")
    command = [
        ffmpeg,
        "-nostdin",
        "-hide_banner",
        "-loglevel",
        "error",
        "-y",
        "-ss",
        f"{start:.3f}",
        "-i",
        str(video_path),
    ]
    if wav_path.is_file():
        command.extend(["-ss", f"{start:.3f}", "-i", str(wav_path)])
    command.extend(
        ["-t", f"{duration:.3f}", "-c:v", "libx264", "-preset", "veryfast"]
    )
    if wav_path.is_file():
        command.extend(
            ["-map", "0:v:0", "-map", "1:a:0", "-c:a", "aac", "-shortest"]
        )
    else:
        command.extend(["-map", "0:v:0", "-map", "0:a?", "-c:a", "aac"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(output_path.stem + ".tmp" + output_path.suffix)
    command.extend(["-movflags", "+faststart", str(temporary)])
    subprocess.run(command, check=True)
    temporary.replace(output_path)


def _contact_sheet(image_paths: Sequence[Path], output_path: Path) -> None:
    import cv2

    images = [cv2.imread(str(path)) for path in image_paths]
    images = [image for image in images if image is not None]
    if not images:
        return
    thumb_width, thumb_height = 480, 270
    thumbnails = [cv2.resize(image, (thumb_width, thumb_height)) for image in images]
    columns = 4
    rows = math.ceil(len(thumbnails) / columns)
    blank = np.full_like(thumbnails[0], 255)
    while len(thumbnails) < rows * columns:
        thumbnails.append(blank.copy())
    sheet_rows = [
        cv2.hconcat(thumbnails[index : index + columns])
        for index in range(0, len(thumbnails), columns)
    ]
    sheet = cv2.vconcat(sheet_rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not cv2.imwrite(str(output_path), sheet):
        raise RuntimeError(f"Failed to save contact sheet: {output_path}")


def _assign_clips(
    pool: pd.DataFrame,
    output_dir: Path,
    context: float,
    event_max_gap: float,
) -> tuple[pd.DataFrame, dict[str, dict[str, Any]]]:
    pool = pool.copy()
    pool["clip_id"] = ""
    pool["clip_path"] = ""
    pool["clip_status"] = ""
    specs: dict[str, dict[str, Any]] = {}

    segmented = pool[pool["segment_id"].fillna("").astype(str) != ""]
    for (episode, segment_id), part in segmented.groupby(["episode", "segment_id"]):
        clip_id = f"{_safe_token(episode)}_SEG_{_safe_token(segment_id)}"
        start = max(0.0, float(part["start_seconds"].min()) - context)
        stop = float(part["stop_seconds"].max()) + context
        relative = Path("media") / "clips" / str(episode) / f"{clip_id}.mp4"
        specs[clip_id] = {
            "episode": str(episode),
            "start": start,
            "stop": stop,
            "path": output_dir / relative,
        }
        pool.loc[part.index, "clip_id"] = clip_id
        pool.loc[part.index, "clip_path"] = relative.as_posix()

    unsegmented = pool[pool["segment_id"].fillna("").astype(str) == ""]
    for episode, episode_rows in unsegmented.groupby("episode"):
        ordered = episode_rows.sort_values(["time", "candidate_id"])
        groups: list[list[int]] = []
        current: list[int] = []
        previous_time: float | None = None
        for row_index, row in ordered.iterrows():
            timestamp = float(row["time"])
            if previous_time is None or timestamp - previous_time <= event_max_gap:
                current.append(row_index)
            else:
                groups.append(current)
                current = [row_index]
            previous_time = timestamp
        if current:
            groups.append(current)
        for indices in groups:
            first_frame = int(pool.loc[indices, "frame_idx"].min())
            clip_id = f"{_safe_token(episode)}_EVT_F{first_frame:07d}"
            start = max(0.0, float(pool.loc[indices, "time"].min()) - context)
            stop = float(pool.loc[indices, "time"].max()) + context
            relative = Path("media") / "clips" / str(episode) / f"{clip_id}.mp4"
            specs[clip_id] = {
                "episode": str(episode),
                "start": start,
                "stop": stop,
                "path": output_dir / relative,
            }
            pool.loc[indices, "clip_id"] = clip_id
            pool.loc[indices, "clip_path"] = relative.as_posix()
    return pool, specs


def _write_candidate_media(
    paths: DatasetPaths,
    pool: pd.DataFrame,
    videos: dict[str, Path],
    overwrite: bool,
) -> pd.DataFrame:
    pool = pool.copy()
    pool["face_crop_path"] = ""
    pool["full_frame_path"] = ""
    pool["media_status"] = "pending"
    readers: dict[str, VideoFrameReader] = {}
    try:
        for row_index, row in pool.iterrows():
            episode = str(row["episode"])
            if episode not in videos:
                raise KeyError(f"{episode} is missing from {paths.video_list}")
            if episode not in readers:
                readers[episode] = VideoFrameReader(videos[episode])
            reader = readers[episode]
            frame = reader.read(int(row["frame_idx"]))
            height, width = frame.shape[:2]
            pool.at[row_index, "frame_width"] = width
            pool.at[row_index, "frame_height"] = height

            candidate_id = str(row["candidate_id"])
            full_relative = (
                Path("media") / "full_frames" / episode / f"{candidate_id}.jpg"
            )
            full_path = paths.output_dir / full_relative
            if overwrite or not full_path.is_file() or full_path.stat().st_size == 0:
                _draw_full_frame(frame, row, full_path)
            pool.at[row_index, "full_frame_path"] = full_relative.as_posix()

            if row["record_type"] == "AVSF" and row["recovery_status"] == "resolved":
                face_relative = (
                    Path("media") / "face_crops" / episode / f"{candidate_id}.jpg"
                )
                face_path = paths.output_dir / face_relative
                if overwrite or not face_path.is_file() or face_path.stat().st_size == 0:
                    _write_face_crop(frame, row, face_path)
                pool.at[row_index, "face_crop_path"] = face_relative.as_posix()
                pool.at[row_index, "media_status"] = "complete"
            elif row["record_type"] == "AVSF":
                pool.at[row_index, "media_status"] = "full_frame_only_unresolved_bbox"
            else:
                pool.at[row_index, "media_status"] = "full_frame_only_no_active_face"
    finally:
        for reader in readers.values():
            reader.close()
    pool["face_position"] = pool.apply(_face_position, axis=1)
    return pool


def _write_all_clips(
    pool: pd.DataFrame,
    specs: dict[str, dict[str, Any]],
    videos: dict[str, Path],
    ffmpeg: str,
    overwrite: bool,
) -> pd.DataFrame:
    executable = shutil.which(ffmpeg) or (str(Path(ffmpeg)) if Path(ffmpeg).is_file() else "")
    if not executable:
        raise FileNotFoundError(f"ffmpeg executable not found: {ffmpeg}")
    statuses: dict[str, str] = {}
    for clip_id, spec in sorted(specs.items()):
        episode = str(spec["episode"])
        if episode not in videos:
            raise KeyError(f"{episode} is missing from the video list")
        video_path = videos[episode]
        wav_path = video_path.with_suffix(".wav")
        if not wav_path.is_file():
            raise FileNotFoundError(
                f"Audio required for review clip is missing: {wav_path}"
            )
        output_path = Path(spec["path"])
        if output_path.is_file() and output_path.stat().st_size > 0 and not overwrite:
            statuses[clip_id] = "existing"
            continue
        _write_clip(
            video_path,
            output_path,
            executable,
            float(spec["start"]),
            float(spec["stop"]),
        )
        statuses[clip_id] = "written"
    result = pool.copy()
    result["clip_status"] = result["clip_id"].map(statuses).fillna("missing")
    return result


def _recommend_candidates(pool: pd.DataFrame, per_class: int) -> pd.DataFrame:
    result = pool.copy()
    result["is_recommended"] = "No"
    result["recommendation_rank"] = ""
    result["recommendation_reason"] = ""

    for label in ["TP", "FP", "FN", "TN"]:
        eligible = result[result["auto_label"] == label].copy()
        if label in {"TP", "FP"}:
            eligible = eligible[
                (eligible["record_type"] == "AVSF")
                & (eligible["recovery_status"] == "resolved")
            ]
        seen: dict[str, set[str]] = defaultdict(set)
        selected_indices: list[int] = []
        selected_units: set[str] = set()

        while len(selected_indices) < per_class:
            choices: list[tuple[float, float, str, int, list[str]]] = []
            for row_index, row in eligible.iterrows():
                if row_index in selected_indices:
                    continue
                unit = _clean_text(row["segment_id"]) or _clean_text(row["clip_id"]) or str(
                    row["candidate_id"]
                )
                if unit in selected_units:
                    continue
                if any(
                    row["episode"] == result.loc[chosen, "episode"]
                    and abs(float(row["time"]) - float(result.loc[chosen, "time"])) < 0.5
                    for chosen in selected_indices
                ):
                    continue

                score = 0.0
                reasons: list[str] = []
                annotated = _clean_text(row["annotated_speaker"])
                mapped = _clean_text(row["mapped_speaker"])
                if annotated and annotated not in seen["annotated_speaker"]:
                    score += 100.0 if label != "FP" else 60.0
                    reasons.append(f"new annotated role: {annotated}")
                if label == "FP" and mapped and mapped not in seen["mapped_speaker"]:
                    score += 60.0
                    reasons.append(f"new mapped role: {mapped}")
                dimensions = [
                    ("episode", 20.0),
                    ("bbox_area_bin", 10.0),
                    ("face_position", 10.0),
                    ("face_orientation", 10.0),
                    ("n_visible_faces", 5.0),
                ]
                for dimension, weight in dimensions:
                    value = _clean_text(row[dimension])
                    if value and value not in seen[dimension]:
                        score += weight
                        reasons.append(f"new {dimension}: {value}")
                purity = float(row["cluster_purity"]) if pd.notna(row["cluster_purity"]) else 0.0
                cosine = float(row["top_cosine"]) if pd.notna(row["top_cosine"]) else 0.0
                duration = max(0.0, float(row["stop_seconds"]) - float(row["start_seconds"]))
                quality = purity * 2.0 + cosine + min(duration, 10.0) / 100.0
                choices.append((score, quality, str(row["candidate_id"]), row_index, reasons))

            if not choices:
                break
            _, _, _, chosen_index, reasons = max(
                choices, key=lambda item: (item[0], item[1], item[2])
            )
            chosen = result.loc[chosen_index]
            selected_indices.append(chosen_index)
            unit = _clean_text(chosen["segment_id"]) or _clean_text(chosen["clip_id"]) or str(
                chosen["candidate_id"]
            )
            selected_units.add(unit)
            for dimension in [
                "annotated_speaker",
                "mapped_speaker",
                "episode",
                "bbox_area_bin",
                "face_position",
                "face_orientation",
                "n_visible_faces",
            ]:
                value = _clean_text(chosen[dimension])
                if value:
                    seen[dimension].add(value)
            rank = len(selected_indices)
            result.at[chosen_index, "is_recommended"] = "Yes"
            result.at[chosen_index, "recommendation_rank"] = rank
            result.at[chosen_index, "recommendation_reason"] = (
                "; ".join(reasons) if reasons else "quality tie-break within covered strata"
            )
        if len(selected_indices) < per_class:
            print(
                f"[WARNING] {label}: recommended {len(selected_indices)} distinct candidates "
                f"(requested {per_class})."
            )
    return result


def _load_manual_reviews(workbook_path: Path) -> dict[str, dict[str, Any]]:
    if not workbook_path.is_file():
        return {}
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if "Candidates" not in workbook.sheetnames:
        raise ValueError(f"Existing workbook has no Candidates sheet: {workbook_path}")
    sheet = workbook["Candidates"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    positions = {header: index for index, header in enumerate(headers)}
    missing = [column for column in ["candidate_id", *MANUAL_REVIEW_COLUMNS] if column not in positions]
    if missing:
        raise ValueError(f"Existing workbook is missing manual-review columns: {missing}")
    reviews: dict[str, dict[str, Any]] = {}
    for values in rows:
        candidate_id = values[positions["candidate_id"]]
        if candidate_id is None:
            continue
        reviews[str(candidate_id)] = {
            column: values[positions[column]] for column in MANUAL_REVIEW_COLUMNS
        }
    workbook.close()
    return reviews


def _excel_value(value: Any) -> Any:
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    if value is None or pd.isna(value):
        return None
    if isinstance(value, float) and not np.isfinite(value):
        return None
    return value


def _write_review_workbook(paths: DatasetPaths, pool: pd.DataFrame) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook_path = paths.review_dir / "candidate_review.xlsx"
    manual_reviews = _load_manual_reviews(workbook_path)
    pool = pool.copy()
    defaults = {
        "final_label": "",
        "review_status": "Unreviewed",
        "selected_for_paper": "No",
        "scenario_tags": "",
        "review_note": "",
    }
    for column, default in defaults.items():
        pool[column] = [
            manual_reviews.get(str(candidate_id), {}).get(column, default) or default
            for candidate_id in pool["candidate_id"]
        ]

    generated_order = [
        "candidate_id",
        "record_type",
        "dataset",
        "episode",
        "segment_id",
        "record_index",
        "time",
        "frame_idx",
        "start_seconds",
        "stop_seconds",
        "annotated_speaker",
        "mapped_speaker",
        "cluster_label",
        "cluster_purity",
        "speaker_face_visible",
        "visible_faces",
        "n_visible_faces",
        "face_orientation",
        "x1",
        "y1",
        "x2",
        "y2",
        "bbox_area",
        "bbox_area_bin",
        "frame_width",
        "frame_height",
        "face_position",
        "recovery_method",
        "recovery_status",
        "top_cosine",
        "cosine_margin",
        "auto_label",
        "auto_reason",
        "is_recommended",
        "recommendation_rank",
        "recommendation_reason",
        "face_crop_path",
        "full_frame_path",
        "clip_id",
        "clip_path",
        "clip_status",
        "media_status",
    ]
    columns = generated_order + MANUAL_REVIEW_COLUMNS

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme_rows = [
        ["EXP-02 Active Visual-Speaker Candidate Review"],
        ["Purpose", "One auditable review table for all AVSF and no-active-segment records."],
        ["auto_label", "Script suggestion only; it is not a TalkNet ground-truth label."],
        ["final_label", "Manual final choice: TP, FP, FN, TN, Ambiguous, or Exclude."],
        ["scenario_tags", "Free-text tags such as multi-person, side-face, occlusion, indoor."],
        ["selected_for_paper", "Set Yes only after reviewing the full-frame clip with audio."],
        ["Media", "Links are relative to this workbook; do not move it outside review/."],
        ["Reruns", "Generated columns refresh by candidate_id; manual columns are preserved."],
    ]
    for row in readme_rows:
        readme.append(row)
    readme["A1"].font = Font(bold=True, size=14)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 95

    sheet = workbook.create_sheet("Candidates")
    sheet.append(columns)
    for _, row in pool[columns].iterrows():
        sheet.append([_excel_value(row[column]) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{sheet.max_row}"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    generated_fill = PatternFill("solid", fgColor="F2F2F2")
    manual_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for column_index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            42, max(12, len(column) + 2)
        )
        fill = manual_fill if column in MANUAL_REVIEW_COLUMNS else generated_fill
        for column_cells in sheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=sheet.max_row,
        ):
            for cell in column_cells:
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    positions = {column: index + 1 for index, column in enumerate(columns)}
    for path_column in ["face_crop_path", "full_frame_path", "clip_path"]:
        column_index = positions[path_column]
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value:
                cell.hyperlink = "../" + str(cell.value)
                cell.style = "Hyperlink"

    validations = [
        ("final_label", FINAL_LABELS),
        ("review_status", ["Unreviewed", "Reviewed", "Needs follow-up"]),
        ("selected_for_paper", ["No", "Yes"]),
    ]
    for column, values in validations:
        validation = DataValidation(
            type="list", formula1='"' + ",".join(values) + '"', allow_blank=True
        )
        sheet.add_data_validation(validation)
        letter = get_column_letter(positions[column])
        validation.add(f"{letter}2:{letter}{max(2, sheet.max_row)}")

    coverage = workbook.create_sheet("Coverage")
    coverage.append(["Final label", "Reviewed rows", "Selected for paper"])
    final_letter = get_column_letter(positions["final_label"])
    selected_letter = get_column_letter(positions["selected_for_paper"])
    speaker_letter = get_column_letter(positions["annotated_speaker"])
    for label in FINAL_LABELS:
        row_number = coverage.max_row + 1
        coverage.append(
            [
                label,
                f'=COUNTIF(Candidates!${final_letter}:${final_letter},A{row_number})',
                (
                    f'=COUNTIFS(Candidates!${final_letter}:${final_letter},A{row_number},'
                    f'Candidates!${selected_letter}:${selected_letter},"Yes")'
                ),
            ]
        )
    coverage.append([])
    coverage.append(["Annotated role", "Manually labelled", "Selected for paper"])
    roles = sorted(
        value for value in {_clean_text(item) for item in pool["annotated_speaker"]} if value
    )
    for role in roles:
        row_number = coverage.max_row + 1
        coverage.append(
            [
                role,
                (
                    f'=COUNTIFS(Candidates!${speaker_letter}:${speaker_letter},A{row_number},'
                    f'Candidates!${final_letter}:${final_letter},"<>")'
                ),
                (
                    f'=COUNTIFS(Candidates!${speaker_letter}:${speaker_letter},A{row_number},'
                    f'Candidates!${selected_letter}:${selected_letter},"Yes")'
                ),
            ]
        )
    for cell in coverage[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    coverage.freeze_panes = "A2"
    coverage.column_dimensions["A"].width = 28
    coverage.column_dimensions["B"].width = 20
    coverage.column_dimensions["C"].width = 20

    paths.review_dir.mkdir(parents=True, exist_ok=True)
    if workbook_path.is_file():
        backup_dir = paths.review_dir / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        shutil.copy2(workbook_path, backup_dir / f"candidate_review_{timestamp}.xlsx")
    temporary = paths.review_dir / "candidate_review.tmp.xlsx"
    workbook.save(temporary)
    temporary.replace(workbook_path)
    return workbook_path


def build_candidates(paths: DatasetPaths, args: argparse.Namespace) -> None:
    annotations = load_annotations(paths)
    identity = load_identity_outputs(paths)
    bbox_csv = data_artifact(paths, "active_face_bbox_records.csv")
    require_files([bbox_csv], "recovered bbox records")
    bbox_frame = pd.read_csv(bbox_csv)
    require_files([paths.subseg_json, paths.video_list], "candidate input")
    subsegments = read_json(paths.subseg_json)
    pool = _candidate_pool(
        paths,
        annotations,
        identity,
        bbox_frame,
        subsegments,
    )
    videos = read_video_list(paths.video_list)
    pool, clip_specs = _assign_clips(
        pool, paths.output_dir, args.clip_context, args.event_max_gap
    )
    pool = _write_candidate_media(paths, pool, videos, args.overwrite)
    if args.write_clips:
        pool = _write_all_clips(pool, clip_specs, videos, args.ffmpeg, args.overwrite)
    else:
        pool["clip_status"] = "skipped"
    pool = _recommend_candidates(pool, args.candidates_per_class)

    paths.data_dir.mkdir(parents=True, exist_ok=True)
    generated_columns = [
        column for column in pool.columns if column not in MANUAL_REVIEW_COLUMNS
    ]
    pool[generated_columns].to_csv(
        paths.data_dir / "candidate_generated_records.csv", index=False
    )
    workbook_path = _write_review_workbook(paths, pool)

    recommended = pool[pool["is_recommended"] == "Yes"].copy()
    for label in ["TP", "FP", "FN", "TN"]:
        part = recommended[recommended["auto_label"] == label].sort_values(
            "recommendation_rank"
        )
        image_paths = [
            paths.output_dir / str(value)
            for value in part["full_frame_path"]
            if _clean_text(value)
        ]
        _contact_sheet(
            image_paths,
            paths.recommendations_dir / "contact_sheets" / f"{label}.jpg",
        )
    counts = Counter(pool["auto_label"].tolist())
    recommendation_counts = Counter(recommended["auto_label"].tolist())
    write_manifest(
        paths,
        "candidates",
        [
            paths.annotation,
            paths.subseg_json,
            paths.video_list,
            bbox_csv,
            data_artifact(paths, "identity_agreement_records.csv"),
        ],
    )
    print(
        f"[INFO] {paths.short_name}: all candidates={dict(counts)}; "
        f"recommended={dict(recommendation_counts)}"
    )
    print(f"[INFO] Review workbook: {workbook_path}")


def run_task(paths: DatasetPaths, args: argparse.Namespace) -> None:
    print(f"[INFO] Dataset: {paths.tv_name}")
    print(f"[INFO] Existing experiment: {paths.exp_dir}")
    print(f"[INFO] EXP-02 output: {paths.output_dir}")
    if args.task in {"all", "summarize"}:
        summarize_dataset(paths)
        if args.require_reference_identity:
            require_reference_identity(paths)
    if args.task in {"all", "recover-bbox"}:
        recover_bboxes(paths, args)
    if args.task in {"all", "plot-bbox"}:
        plot_bbox_histogram(paths, args.min_recovery_rate)
    if args.task in {"all", "candidates"}:
        if not data_artifact(paths, "identity_agreement_records.csv").is_file():
            summarize_dataset(paths)
        build_candidates(paths, args)


def main() -> None:
    args = parse_args()
    for tv_name in args.tv_name:
        run_task(resolve_paths(args, tv_name), args)


if __name__ == "__main__":
    main()
